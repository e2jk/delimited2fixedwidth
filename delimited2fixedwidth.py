#!/usr/bin/env python
# -*- coding: utf-8 -*-

#    This file is part of delimited2fixedwidth and is MIT-licensed.

import sys
import argparse
import logging
import os
import csv
from openpyxl import load_workbook
import re
import datetime


def write_output_file(output_content, output_file):
    with open(output_file, "w") as ofile:
        ofile.write("\n".join(output_content))


def pad_output_value(val, output_format, length):
    if output_format in ("Integer", "Decimal"):
        # Numbers get padded with 0's added in front (to the left)
        val = str(val).zfill(length)
    else:
        # Strings get padded with spaces added to the right
        format_template = "{:<%d}" % length
        val = format_template.format(val)
    return val


def convert_cell(value, output_format, idx_col, idx_row):
    converted_value = ""
    if "Time" == output_format:
        m = re.match(r"(\d{2})(:)?(\d{2})", value)
        if m:
            converted_value = "%s%s" % (m.group(1), m.group(3))
        else:
            logging.critical(
                "Invalid time format '%s' in field %d on row %d (ignoring "
                "the header). Exiting..." % (value, idx_col, idx_row)
            )
            sys.exit(17)
    elif output_format in (
        "Date (DD/MM/YYYY to YYYYMMDD)",
        "Date (MM/DD/YYYY to YYYYMMDD)",
    ):
        if "Date (DD/MM/YYYY to YYYYMMDD)" == output_format:
            m = re.match(r"([0123]?\d)/([01]?\d)/(\d{4})", value)
            if m:
                year = m.group(3)
                month = m.group(2).zfill(2)
                day = m.group(1).zfill(2)
        elif "Date (MM/DD/YYYY to YYYYMMDD)" == output_format:
            m = re.match(r"([01]?\d)/([0123]?\d)/(\d{4})", value)
            if m:
                year = m.group(3)
                month = m.group(1).zfill(2)
                day = m.group(2).zfill(2)
        if m:
            converted_value = "%s%s%s" % (year, month, day)
            # Is it a valid date?
            try:
                datetime.datetime.strptime(converted_value, "%Y%m%d")
            except ValueError:
                converted_value = ""
        if not converted_value:
            logging.critical(
                "Invalid date value '%s' for format '%s' in field %d on row "
                "%d (ignoring the header). Exiting..."
                % (value, output_format, idx_col, idx_row)
            )
            sys.exit(18)
    elif "Decimal" == output_format:
        # Decimal numbers must be sent with 2 decimal places and
        # *without* the decimal separator
        try:
            # Convert to float, multiply by 100, round without decimals,
            # convert to integer (to drop the extra decimal values) then
            # finally back to string...
            converted_value = float(value)
            converted_value = converted_value * 100
            converted_value = round(converted_value, 0)
            converted_value = int(converted_value)
            converted_value = str(converted_value)
        except ValueError:
            logging.critical(
                "Invalid decimal format '%s' in field %d on row %d (ignoring "
                "the header). Exiting..." % (value, idx_col, idx_row)
            )
            sys.exit(19)
    else:
        converted_value = value
    return converted_value


def convert_content(input_content, config, date_field_to_report_on=None):
    output_content = []
    if date_field_to_report_on:
        # Argument is 1-based
        date_field_to_report_on -= 1
    oldest_date = "99999999"
    most_recent_date = "00000000"
    for idx_row, row in enumerate(input_content):
        converted_row_content = []
        # Confirm that the input_content doesn't have more fields than are
        # defined in the configuration file
        if len(row) > len(config):
            logging.critical(
                "Row %d (ignoring the header) has more fields than are "
                "defined in the configuration file! The row has %d fields "
                "while the configuration defines only %d possible fields. "
                "Exiting..." % (idx_row + 1, len(row), len(config))
            )
            sys.exit(23)
        for idx_col, cell in enumerate(row):
            output_format = config[idx_col]["output_format"]
            length = config[idx_col]["length"]

            if config[idx_col]["skip_field"]:
                cell = ""
            else:
                cell = convert_cell(cell, output_format, idx_col + 1, idx_row + 1)

            # Confirm that the length of the field (before padding) is less
            # than the maximum allowed length
            if len(cell) > config[idx_col]["length"]:
                logging.critical(
                    "Field %d on row %d (ignoring the header) is too long! "
                    "Length: %d, max length %d. Exiting..."
                    % (idx_col + 1, idx_row + 1, len(cell), config[idx_col]["length"])
                )
                sys.exit(20)

            padded_output_value = pad_output_value(cell, output_format, length)
            converted_row_content.append(padded_output_value)

            if date_field_to_report_on and date_field_to_report_on == idx_col:
                if padded_output_value < oldest_date:
                    oldest_date = padded_output_value
                if padded_output_value > most_recent_date:
                    most_recent_date = padded_output_value
        # Process fields not in the input content but defined in the
        # configuration file: empty padding, based on the defined output format
        for idx_col in range(len(row), len(config)):
            output_format = config[idx_col]["output_format"]
            length = config[idx_col]["length"]
            padded_output_value = pad_output_value("", output_format, length)
            converted_row_content.append(padded_output_value)
        output_content.append("".join(converted_row_content))

    logging.debug("The output content:\n%s" % "\n".join(output_content))
    return (output_content, oldest_date, most_recent_date)


def read_input_file(input_file, delimiter, quotechar, skip_header, skip_footer):
    content = None
    with open(input_file, newline="") as csvfile:
        content = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
        # Skip the header and footer if necessary
        content = list(content)
        num_lines = len(content)
        content = content[skip_header : num_lines - skip_footer]
        logging.debug(
            "There are %d lines in the input file %s:" % (num_lines, input_file)
        )
        if skip_header > 0 or skip_footer > 0:
            logging.debug(
                "Skipping %d header and %d footer lines" % (skip_header, skip_footer)
            )
        for row in content:
            logging.debug(" ||| ".join(row))
    return content


def load_config(config_file):
    config = []
    supported_output_formats = (
        "Integer",
        "Decimal",
        "Time",
        "Text",
        "Date (DD/MM/YYYY to YYYYMMDD)",
    )
    supported_skip_field = ("True", "False", "", None)
    logging.debug("Loading configuration %s" % config_file)

    # Open the configuration file (an Excel .xlsx file)
    wk = load_workbook(filename=config_file)
    ws = wk.active  # Get active worksheet or wk['some_worksheet']

    # Analyze the header to identify the relevant columns
    length_col = -1
    output_format_col = -1
    skip_field_col = -1
    for row in ws.iter_rows(max_row=1):
        for idx, cell in enumerate(row):
            if "Length" == cell.value:
                length_col = idx
            elif "Output format" == cell.value:
                output_format_col = idx
            elif "Skip field" == cell.value:
                skip_field_col = idx
    column_indices = (length_col, output_format_col, skip_field_col)
    if -1 in column_indices:
        logging.critical(
            "Invalid config file, missing one of the columns 'Length', "
            "'Output format' or 'Skip field'. Exiting..."
        )
        sys.exit(13)

    # Loop over all the config rows (skipping the header)
    for idx_row, row in enumerate(ws.iter_rows(min_row=2)):
        config.append({})
        for idx_col, cell in enumerate(row):
            if idx_col == length_col:
                config[idx_row]["length"] = -1
                if isinstance(cell.value, int):
                    config[idx_row]["length"] = cell.value
                if isinstance(cell.value, str) and cell.value.isnumeric():
                    config[idx_row]["length"] = int(cell.value)
                if config[idx_row]["length"] < 0:
                    logging.critical(
                        "Invalid value '%s' for the 'Length' column on row "
                        "%d, must be a positive number. Exiting..."
                        % (cell.value, idx_row + 2)
                    )
                    sys.exit(14)
            if idx_col == output_format_col:
                if cell.value in supported_output_formats:
                    config[idx_row]["output_format"] = cell.value
                else:
                    logging.critical(
                        "Invalid output format '%s' on row %d, must be one  "
                        "of '%s'. Exiting..."
                        % (
                            cell.value,
                            idx_row + 2,
                            "', '".join(supported_output_formats),
                        )
                    )
                    sys.exit(15)
            if idx_col == skip_field_col:
                if cell.value in supported_skip_field:
                    config[idx_row]["skip_field"] = "True" == cell.value
                else:
                    logging.critical(
                        "Invalid value '%s' for the 'Skip field' column on "
                        "row %d, must be one  of 'True', 'False' or empty. "
                        "Exiting..." % (cell.value, idx_row + 2)
                    )
                    sys.exit(16)

    logging.info("Config '%s' loaded successfully" % config_file)
    logging.debug(config)
    return config


def get_version(rel_path):
    with open(rel_path) as f:
        for line in f.read().splitlines():
            if line.startswith("__version__"):
                return line.split('"')[1]
        else:
            raise RuntimeError("Unable to find version string.")


def validate_shared_args(args):
    if not os.path.isfile(args.input):
        logging.critical("The specified input file does not exist. Exiting...")
        sys.exit(10)
    if not os.path.isfile(args.config):
        logging.critical("The specified configuration file does not exist. Exiting...")
        sys.exit(12)
    if args.skip_header != 0:
        try:
            args.skip_header = int(args.skip_header)
        except ValueError:
            logging.critical("The `--skip-header` argument must be numeric. Exiting...")
            sys.exit(21)
    if args.skip_footer != 0:
        try:
            args.skip_footer = int(args.skip_footer)
        except ValueError:
            logging.critical("The `--skip-footer` argument must be numeric. Exiting...")
            sys.exit(22)


def add_shared_args(parser):
    parser.add_argument(
        "-i", "--input", help="Specify the input file", action="store", required=True
    )
    parser.add_argument(
        "-c",
        "--config",
        help="Specify the configuration file",
        action="store",
        required=True,
    )
    parser.add_argument(
        "-dl",
        "--delimiter",
        help="The field delimiter used in the input file (default ,)",
        action="store",
        required=False,
        default=",",
    )
    parser.add_argument(
        "-q",
        "--quotechar",
        help="The character used to wrap textual fields in the input file "
        '(default ")',
        action="store",
        required=False,
        default='"',
    )
    parser.add_argument(
        "-sh",
        "--skip-header",
        help="The number of header lines to skip (default 0)",
        action="store",
        required=False,
        default=0,
    )
    parser.add_argument(
        "-sf",
        "--skip-footer",
        help="The number of footer lines to skip (default 0)",
        action="store",
        required=False,
        default=0,
    )


def parse_args(arguments):
    parser = argparse.ArgumentParser(
        description="Convert files from delimited (e.g. CSV) to fixed width " "format"
    )
    parser.add_argument(
        "--version",
        action="version",
        version="%s %s" % ("%(prog)s", get_version("__init__.py")),
    )

    parser.add_argument(
        "-o", "--output", help="Specify the output file", action="store", required=True
    )
    parser.add_argument(
        "-x",
        "--overwrite-file",
        help="Allow to overwrite the output file",
        action="store_true",
        required=False,
    )

    add_shared_args(parser)

    parser.add_argument(
        "-d",
        "--debug",
        help="Print lots of debugging statements",
        action="store_const",
        dest="loglevel",
        const=logging.DEBUG,
        default=logging.WARNING,
    )
    parser.add_argument(
        "-v",
        "--verbose",
        help="Be verbose",
        action="store_const",
        dest="loglevel",
        const=logging.INFO,
    )
    args = parser.parse_args(arguments)

    # Configure logging level
    if args.loglevel:
        logging.basicConfig(level=args.loglevel)
        args.logging_level = logging.getLevelName(args.loglevel)

    # Validate if the arguments are used correctly
    if os.path.isfile(args.output) and not args.overwrite_file:
        logging.critical(
            "The specified output file does already exist, will "
            "NOT overwrite. Add the `--overwrite-file` argument "
            "to allow overwriting. Exiting..."
        )
        sys.exit(11)
    validate_shared_args(args)

    logging.debug("These are the parsed arguments:\n'%s'" % args)
    return args


def process(
    input,
    output,
    config,
    delimiter,
    quotechar,
    skip_header,
    skip_footer,
    date_field_to_report_on=None,
):
    config = load_config(config)

    input_content = read_input_file(
        input, delimiter, quotechar, skip_header, skip_footer
    )

    (output_content, oldest_date, most_recent_date) = convert_content(
        input_content, config, date_field_to_report_on
    )

    write_output_file(output_content, output)

    return (len(input_content), oldest_date, most_recent_date)


def init():
    if __name__ == "__main__":
        # Parse the provided command-line arguments
        args = parse_args(sys.argv[1:])

        process(
            args.input,
            args.output,
            args.config,
            args.delimiter,
            args.quotechar,
            args.skip_header,
            args.skip_footer,
        )


init()
